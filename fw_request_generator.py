import pandas as pd
from openpyxl import load_workbook

inventory_file = "/Users/girlpower/Desktop/inventory_file.xlsx"
template_file = "/Users/girlpower/Desktop/template_file.xlsx"
storage_file = "/Users/girlpower/Desktop/storage_file.xlsx"

inventory_df = pd.read_excel(inventory_file)
template_df = pd.read_excel(template_file)
storage_df = pd.read_excel(storage_file)

region = input("Please, enter the region ASIA/ME/NE/SE : ")

print("")
print("   - AXA Banque")
print("   - AXA Belgium")
print("   - AXA France")
print("   - AXA Go")
print("   - AXA Italy")
print("   - AXA Life Japan")
print("   - AXA Partners")
print("   - AXA Spain")
print("   - AXA Uk")

entity = input("Please COPY/PAST the entity : ")

exposure = input("\nPlease enter the exposure Internal/External: ")
provider_pod = input("\nPlease confirm the use of POD filter by typing POD : ")
provider = input("\nPlease, enter the provider Azure/AWS: ")

core_filter = inventory_df[(inventory_df['Type'] == 'Core') & (inventory_df['Region'] == region)]
core_ip = core_filter['IP'].values
core_name = core_filter['Name Server'].values

print("\nCORE IP: ", core_ip)
print("Core Name: ", core_name)


pod_proxy_filter = inventory_df[(inventory_df['Type'] == 'Proxy') & (inventory_df['Provider'] == provider_pod)]
pod_proxy_ip = pod_proxy_filter['IP'].values
pod_proxy_name = pod_proxy_filter['Name Server'].values

print("\nPod PROXY IP: ", pod_proxy_ip)
print("Pod PROXY Name: ", pod_proxy_name)


mpi_proxy_filter = inventory_df[(inventory_df['Type'] == 'Proxy') & (inventory_df['Provider'] == provider)]
mpi_proxy_ip = mpi_proxy_filter['IP'].values
mpi_proxy_name = mpi_proxy_filter['Name Server'].values

print("\nMPI PROXY IP: ", mpi_proxy_ip)
print("MPI PROXY Name: ", mpi_proxy_name)

storage_filter = storage_df[(storage_df['Entity'] == entity ) & (storage_df['EXPOSURE'] == exposure) & (storage_df['Provider'] == provider)]
storage_subnet = storage_filter['Subnet'].values
storage_name = storage_filter['Storage Name'].values

print("\n ============= Flow opening Options ============= ")
print("1 - Infrastructure flow opening")
print("2 - Flow between MPI Proxies & CORE")
print("3 - Flow between CORE, MPI Proxies, POD Proxies & Source NAS")

choice = int(input("Your choice please : "))
print("")

if choice == 1:
    row_index = 2 

    for i in range(len(pod_proxy_ip)):
        while row_index >= len(template_df):
            
            template_df.loc[len(template_df)] = [None] * len(template_df.columns)

        template_df.iloc[row_index, 0] = pod_proxy_ip[i]      
        template_df.iloc[row_index, 1] = pod_proxy_name[i]    
        template_df.iloc[row_index, 2] = core_ip[0]           
        template_df.iloc[row_index, 3] = core_name[0]           
        template_df.iloc[row_index, 4] = "tcp/2049,tcp/635,tcp/111,tcp/445"
        template_df.iloc[row_index, 5] = "any"
        template_df.iloc[row_index, 6] = "Allow"
        template_df.iloc[row_index, 7] = "Flow to be opened between POD proxies and the CORE"
        template_df.iloc[row_index, 8] = "Internal (I)"
        template_df.iloc[row_index, 9] = "Internal (I)"        

        row_index += 1

    for i in range(len(mpi_proxy_ip)):
        while row_index >= len(template_df):
            template_df.loc[len(template_df)] = [None] * len(template_df.columns)

        template_df.iloc[row_index, 0] = mpi_proxy_ip[i]      
        template_df.iloc[row_index, 1] = mpi_proxy_name[i]  
        template_df.iloc[row_index, 2] = core_ip[0]                 
        template_df.iloc[row_index, 3] = core_name[0]             
        template_df.iloc[row_index, 4] = "tcp/1970,tcp/1971,tcp/8443"
        template_df.iloc[row_index, 5] = "any"
        template_df.iloc[row_index, 6] = "Allow"
        template_df.iloc[row_index, 7] = "Flow to be opened between MPI proxies and the CORE"
        template_df.iloc[row_index, 8] = "Internal (I)"
        template_df.iloc[row_index, 9] = "Internal (I)"

        row_index += 1
    
    for i in range(len(storage_subnet)):
        while row_index >= len(template_df):
            template_df.loc[len(template_df)] = [None] * len(template_df.columns)

        template_df.iloc[row_index, 2] = storage_subnet[i]
        template_df.iloc[row_index, 3] = storage_name[i]
        template_df.iloc[row_index, 0] = core_ip[0]
        template_df.iloc[row_index, 1] = core_name[0]
        template_df.iloc[row_index, 4] = "tcp/443"
        template_df.iloc[row_index, 5] = "any"
        template_df.iloc[row_index, 6] = "Allow"
        template_df.iloc[row_index, 7] = "Flow to be opened between the CORE & the MPI storage"
        template_df.iloc[row_index, 8] = "Internal (I)"
        template_df.iloc[row_index, 9] = "Internal (I)"
        row_index += 1
 
    template_df.to_excel('/Users/girlpower/Desktop/template_MAJ.xlsx', index=False)

    workbook = load_workbook('/Users/girlpower/Desktop/template_MAJ.xlsx')
    sheet = workbook.active
    sheet.merge_cells('A1:J1')
    workbook.save('/Users/girlpower/Desktop/template_MAJ.xlsx')
    print("Flow request generated. Good luck")

else:
    row_index = 2
    nbr_source_nas = int(input("How many source IPs you want to open flow for them : "))

    source_nas_ips = []
    source_nas_names = []

    for i in range(nbr_source_nas):
        source_nas_ip = input(f"Please, type the source NAS IP for source {i+1}: ")
        source_nas_name = input(f"Please, type the source NAS name for source {i+1}: ")
    
        source_nas_ips.append(source_nas_ip)
        source_nas_names.append(source_nas_name)

    for i in range(len(pod_proxy_ip)):
        for j in range(nbr_source_nas):  
            while row_index >= len(template_df):
                template_df.loc[len(template_df)] = [None] * len(template_df.columns)

            template_df.iloc[row_index, 0] = pod_proxy_ip[i]
            template_df.iloc[row_index, 1] = pod_proxy_name[i]
            template_df.iloc[row_index, 2] = source_nas_ips[j]
            template_df.iloc[row_index, 3] = source_nas_names[j]
            template_df.iloc[row_index, 4] = "tcp/2049,tcp/635,tcp/111,tcp/445"
            template_df.iloc[row_index, 5] = "any"
            template_df.iloc[row_index, 6] = "Allow"
            template_df.iloc[row_index, 7] = "Flow to be opened between POD proxies and NAS"
            template_df.iloc[row_index, 8] = "Internal (I)"
            template_df.iloc[row_index, 9] = "Internal (I)"

            row_index += 1

    for i in range(len(mpi_proxy_ip)):
        for j in range(nbr_source_nas):  
            while row_index >= len(template_df):
                template_df.loc[len(template_df)] = [None] * len(template_df.columns)

            template_df.iloc[row_index, 0] = mpi_proxy_ip[i]
            template_df.iloc[row_index, 1] = mpi_proxy_name[i]
            template_df.iloc[row_index, 2] = source_nas_ips[j]
            template_df.iloc[row_index, 3] = source_nas_names[j]
            template_df.iloc[row_index, 4] = "tcp/1970,tcp/1971,tcp/8443"
            template_df.iloc[row_index, 5] = "any"
            template_df.iloc[row_index, 6] = "Allow"
            template_df.iloc[row_index, 7] = "Flow to be opened between MPI Proxies & source NAS"
            template_df.iloc[row_index, 8] = "Internal (I)"
            template_df.iloc[row_index, 9] = "Internal (I)"

            row_index += 1
    for i in range(nbr_source_nas):
        while row_index >= len(template_df):
            template_df.loc[len(template_df)] = [None] * len(template_df.columns)

        template_df.iloc[row_index, 0] = core_ip[0]
        template_df.iloc[row_index, 1] = core_name[0]
        template_df.iloc[row_index, 2] = source_nas_ips[j]
        template_df.iloc[row_index, 3] = source_nas_names[j]
        template_df.iloc[row_index, 4] = "tcp/443"
        template_df.iloc[row_index, 5] = "any"
        template_df.iloc[row_index, 6] = "Allow"
        template_df.iloc[row_index, 7] = "Flow to be opened between the CORE and the source NAS"
        template_df.iloc[row_index, 8] = "Internal (I)"
        template_df.iloc[row_index, 9] = "Internal (I)"

    template_df.to_excel('/Users/girlpower/Desktop/template_MAJ.xlsx', index=False)
    print("Le template a été mis à jour avec les informations des Pod Proxies et du Core.")